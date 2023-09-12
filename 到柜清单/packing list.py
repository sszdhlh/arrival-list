import csv
import openpyxl
import datetime
import sys
from openpyxl.styles import PatternFill, Font

colory = ["ffeb9c", "9c6500"]
filly = PatternFill('solid', fgColor=colory[0])

colorr = ["cd3131", "006100"]
fillr = PatternFill('solid', fgColor=colorr[0])
fontr = Font('Arial', color = colory[1], size=12)

parentlist=[]    
uselist=[]    
acclist=['COMPACC','DSBATS','TT-BATS-PRIME-A03','DSNET','RAINBOW-NET']
#到cin7里的BOM Master里导出
with open('BOMExport.csv','r',encoding='utf-8') as f:
    file = csv.reader(f)
    next(file)
    parent = ""
    for line in file:
        if len(line) < 2 or line[1] == None or line[1] == '':
            continue
        if line[4] == 'Parent':
            parent = line[1]
        elif line[1] not in acclist:
            parentlist.append(parent)
            uselist.append(line[1])

presale_sku = []
presale_date_list = []
with open("预售.csv", "r", encoding="utf-8") as f:
    file = csv.reader(f)
    next(file)
    for line in file:
        presale_sku.append(line[0].strip())
        date = line[1]
        start_date = ""
        try:
            start_date = datetime.datetime.strptime(date, '%d/%m/%Y').date()
        except:
            start_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        presale_date_list.append(start_date)

        # sys.exit()
    
    
sku_list = []
qty_list = []
sku_order_dict = {}

workbook = openpyxl.load_workbook("EITU0368228.xlsx")
contentsheet = workbook[(workbook.sheetnames[0])]
contentrow = contentsheet.max_row #总行数    

for i in range(2,contentrow):
    if contentsheet.cell(row = i+1, column = 1).value == None:
        break
    sku=contentsheet.cell(row = i+1, column = 1).value.upper().strip()
    qty=contentsheet.cell(row = i+1, column = 4).value

    if sku in presale_sku:
        sku_list.append(sku)
        qty_list.append(int(qty))
        sku_order_dict[sku] = {
            "SYD": "",
            "MEL": "",
            "BNE": "",
            "QTY": [0,0,0],
            "Time": []
        }
    elif sku in uselist:
        # print(sku)
        parent_index = uselist.index(sku)
        parent = parentlist[parent_index]
        if parent in presale_sku and parent not in sku_list:
            sku_list.append(parent)
            qty_list.append(int(qty))
            sku_order_dict[parent] = {
                "SYD": "",
                "MEL": "",
                "BNE": "",
                "QTY": [0,0,0],
                "Time": []
            }

print(sku_list)
print(qty_list)


order_num = []
#导出最近一年的所有订单，选create date， Invoice date， Item Bom load，branch
with open("Cin7 data/1.csv", "r", encoding="utf-8") as f:
    file = csv.reader(f)
    next(file)
    for line in file:
        # if line[8] == 'Use':
        #     continue
        if line[6] == 'SHIPPING' or line[6] == 'OC' or line[6] == 'Installation':
            continue
        
        order_number = line[0].strip()
        sku = line[6].strip()
        branch = line[5]
        date = line[4]
        date = date.split(" ")
        date = date[:3]
        date_str = date[0]+"-"+date[1]+"-"+date[2]
        order_date = datetime.datetime.strptime(date_str, '%d-%b-%Y').date()
        
        
        # if sku.strip() == "SCHAIR-TB-XLHORIZON-WHT":
        if sku in sku_list:
            index = presale_sku.index(sku)
            start_date = presale_date_list[index]
            # print(start_date)
            # sys.exit()
            if order_date >= start_date:
                if "NSW" in branch:
                    sku_order_dict[sku]["SYD"] += order_number + ";\n"
                    sku_order_dict[sku]["QTY"][0] += int(line[7])
                    sku_order_dict[sku]["Time"].append("SYD: " + str(line[7]))
                elif "QLD" in branch:
                    sku_order_dict[sku]["BNE"] += order_number + ";\n"
                    sku_order_dict[sku]["QTY"][2] += int(line[7])
                    sku_order_dict[sku]["Time"].append("BNE: "+ str(line[7]))
                else:
                    sku_order_dict[sku]["MEL"] += order_number + ";\n"
                    sku_order_dict[sku]["QTY"][1] += int(line[7])
                    sku_order_dict[sku]["Time"].append("MEL: "+ str(line[7]))
                
order_list = []
real_list = []
for i in range(len(sku_list)):
    sku = sku_list[i]
    order_list.append(sku_order_dict[sku]["QTY"])

overload_item = []
for i in range(len(sku_list)):
    sku = sku_list[i]
    ratio = sku_order_dict[sku]["QTY"]
    total = sum(ratio)
    
    # if sku == "SCHAIR-TB-XLHORIZON-WHT":
    
    quantity = qty_list[i]
    sku_order_dict[sku]["Time"].reverse()
    order_by_time = sku_order_dict[sku]["Time"]

    if total == 0:
        real_list.append([quantity,0,0])
    else:
        syd_real = 0
        mel_real = 0
        bne_real = 0

        for item in order_by_time:
            branch = item.split(":")[0]
            num = int(item.split(":")[1])
            if branch == "SYD":
                syd_real += num
            elif branch == "MEL":
                mel_real += num
            else:
                bne_real += num
            if syd_real + mel_real + bne_real == quantity:
                break
            elif syd_real + mel_real + bne_real > quantity:
                overload_item.append(sku)
                break

        if syd_real + mel_real + bne_real < quantity:
            rest = quantity - (syd_real + mel_real + bne_real)
            syd_real += rest
                
        result_list = [syd_real,mel_real,bne_real]        
        real_list.append(result_list)



contentsheet.cell(row = 2, column = 6).value = "SYD"    
contentsheet.cell(row = 2, column = 7).value = "MEL" 
contentsheet.cell(row = 2, column = 8).value = "BNE" 

contentsheet.cell(row = 2, column = 9).value = "SYD 总共定单"
contentsheet.cell(row = 2, column = 10).value = "MEL 总共定单"
contentsheet.cell(row = 2, column = 11).value = "BNE 总共定单"
contentsheet.cell(row = 2, column = 12).value = "SYD 欠单"
contentsheet.cell(row = 2, column = 13).value = "MEL 欠单"
contentsheet.cell(row = 2, column = 14).value = "BNE 欠单"

for i in range(2,contentrow):
    if contentsheet.cell(row = i+1, column = 1).value == None:
        break
    sku=contentsheet.cell(row = i+1, column = 1).value.upper().strip()
    if sku in uselist and sku not in presale_sku:
        parent_index = uselist.index(sku)
        parent = parentlist[parent_index]
        sku = parent
    if sku in sku_list:
        for u in range(1, 9):
            if sku not in overload_item:
                contentsheet.cell(row = i+1, column=u).fill = filly
            else:
                contentsheet.cell(row = i+1, column=u).fill = fillr
        
        index = sku_list.index(sku)
        if int(sum(real_list[index])) != int(contentsheet.cell(row = i+1, column = 4).value):
            contentsheet.cell(row = i+1, column = 6).fill = fillr
            contentsheet.cell(row = i+1, column = 7).fill = fillr
            contentsheet.cell(row = i+1, column = 8).fill = fillr

        contentsheet.cell(row = i+1, column = 6).value = str(real_list[index][0])
        contentsheet.cell(row = i+1, column = 7).value = str(real_list[index][1])
        contentsheet.cell(row = i+1, column = 8).value = str(real_list[index][2])
        
        contentsheet.cell(row = i+1, column = 9).value = str(sku_order_dict[sku]["QTY"][0])
        contentsheet.cell(row = i+1, column = 10).value = str(sku_order_dict[sku]["QTY"][1])
        contentsheet.cell(row = i+1, column = 11).value = str(sku_order_dict[sku]["QTY"][2])
        contentsheet.cell(row = i+1, column = 12).value = str(sku_order_dict[sku]["SYD"])
        contentsheet.cell(row = i+1, column = 13).value = str(sku_order_dict[sku]["MEL"])
        contentsheet.cell(row = i+1, column = 14).value = str(sku_order_dict[sku]["BNE"])
        

workbook.save("inbound lists presale.xlsx")
